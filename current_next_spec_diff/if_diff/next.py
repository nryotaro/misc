#!/usr/bin/env python
# -*- coding: utf-8 -*-

def write_file_xls(sheet):
    msg_name = sheet.cell(12,2).value
    assert(sheet.cell(23,4).value == 'cell str val')
    tuples = [(sheet.cell(r,4),sheet.cell(r, 5)) for r in range(24, sheet.nrows)]
    with open (msg_name + '_next.txt', mode = 'w') as f:
        last_p=""
        for r in filter(lambda t: not (len(t[0].value) == 0 and len(t[1].value) == 0), tuples):
            last_p = r[0].value if len(r[0].value) != 0 else last_p
            f.write(last_p + ',' + r[1].value + '\n')

if __name__ == '__main__':
    import sys,csv, xlrd, glob
    from openpyxl import load_workbook

    for f in [f for f in glob.glob("filename")]:
        wb = xlrd.open_workbook(f)
        write_file_xls(wb.sheet_by_index(0))
