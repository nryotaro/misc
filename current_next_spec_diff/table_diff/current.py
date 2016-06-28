#!/usr/bin/env python
# -*- coding: utf-8 -*-

class Col:
    def __init__(self, name, typ, siz):
        self.name = name
        self.typ = typ
        self.siz = siz

def write_file(sheet):
    table_name = sheet.cell(4,2).value
    tuples = [(sheet.cell(r,1),sheet.cell(r, 6), sheet.cell(r,7)) for r in range(8, sheet.nrows)]
    with open (table_name+ '_' + sheet.cell(0,2).value + '_current.txt', mode = 'w') as f:
        for r in filter(lambda t: t[0] != None, tuples):
            siz = str(int(round(r[2].value))) if type(r[2].value) == float else r[2].value
            f.write(r[0].value.lower() + ',' + r[1].value + ',' + siz + '\n')

if __name__ == '__main__':
    import sys,csv, glob,re, xlrd
    from openpyxl import load_workbook

    files=[f for f in glob.glob("current\\db_design\\*.xls")]
    files=list(filter(lambda l: re.match(r".*table_list\)\.xls$",l)== None,files))
    for f in files:
        wb = xlrd.open_workbook(f)
        write_file(wb.sheet_by_index(0))
