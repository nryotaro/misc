#!/usr/bin/env python
# -*- coding: utf-8 -*-

class Col:
    def __init__(self, name, typ, siz):
        self.name = name
        self.typ = typ
        self.siz = siz

def write_file(sheet):
    table_name = sheet['C7'].value.lower()
    tuples = [(sheet['C' + str(r)].value,
      sheet['F' + str(r)].value,
      sheet['G' + str(r)].value) for r in range(11, sheet.max_row+1)]

    with open (table_name  +'_' + sheet['C4'].value +'_next.txt', mode = 'w') as f:
        for r in filter(lambda t: t[0] != None, tuples):
            siz = "" if r[2] == None else str(r[2])
            f.write(r[0].lower() + ',' + r[1] + ',' + siz + '\n')

if __name__ == '__main__':
    import sys,csv
    from openpyxl import load_workbook

    wb = load_workbook(sys.argv[1])
    #sheet = wb.get_sheet_by_name(sys.argv[2])

    for sheet in filter(lambda x: x != 'spec', wb.get_sheet_names()):
        write_file(wb[sheet])
